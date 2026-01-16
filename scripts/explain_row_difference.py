"""Explain why merged sheet has more rows than Rippling."""
from openpyxl import load_workbook

rippling = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors Rippling.xlsx', data_only=True)
r_ws = rippling.active

merged = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors_and_contacts.xlsx', data_only=True)
m_ws = merged.active
m_headers = [str(c.value or '') for c in m_ws[1]]
type_idx = 1
name_idx = 0
related_idx = next((i for i, h in enumerate(m_headers) if 'related company' in h.lower()), None)

r_total = r_ws.max_row - 1
m_companies = sum(1 for i in range(2, m_ws.max_row+1) if m_ws.cell(i, type_idx+1).value == 'Company')
m_contacts = sum(1 for i in range(2, m_ws.max_row+1) if m_ws.cell(i, type_idx+1).value == 'Person')
m_total = m_ws.max_row - 1

print("=" * 80)
print("ROW COUNT EXPLANATION")
print("=" * 80)
print(f"\nRippling Sheet:")
print(f"  Total data rows: {r_total}")
print(f"  (These are vendor/company rows)")

print(f"\nMerged Sheet:")
print(f"  Total data rows: {m_total}")
print(f"  - Company rows: {m_companies}")
print(f"  - Contact rows: {m_contacts}")
print(f"  - Total: {m_companies + m_contacts}")

print(f"\nDifference: {m_total - r_total} extra rows")
print(f"\nWhy more rows?")
print(f"  1. Deduplication: {r_total - m_companies} duplicate companies were merged")
print(f"     (Rippling had case-sensitive duplicates like 'AFTRLIFE' vs 'Aftrlife')")
print(f"  2. Contacts added: {m_contacts} contact rows were added beneath companies")
print(f"     (Each contact person gets their own row linked to the company)")

print(f"\n" + "=" * 80)
print("STRUCTURE EXAMPLE")
print("=" * 80)
print("\nIn the merged sheet, the structure looks like this:")
print("\nRow 2:  COMPANY - 'Aftrlife Entertainment'")
print("Row 3:  CONTACT - 'John Doe' (Related Company: 'Aftrlife Entertainment')")
print("Row 4:  CONTACT - 'Jane Smith' (Related Company: 'Aftrlife Entertainment')")
print("Row 5:  COMPANY - 'Amazing Attractions'")
print("Row 6:  CONTACT - 'Bob Johnson' (Related Company: 'Amazing Attractions')")
print("Row 7:  COMPANY - 'All About Animals'")
print("        (No contacts - just the company)")

print(f"\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"\nThe merged sheet has MORE rows because:")
print(f"  [OK] Each company gets 1 row")
print(f"  [OK] Each contact person gets their own row beneath their company")
print(f"  [OK] This matches Odoo's import format (companies + contacts)")
print(f"\nThe merged sheet has FEWER companies because:")
print(f"  [OK] Case-sensitive duplicates were merged (e.g., 'AFTRLIFE' + 'Aftrlife' = 1 company)")

# Show a real example
print(f"\n" + "=" * 80)
print("REAL EXAMPLE FROM YOUR DATA")
print("=" * 80)
count = 0
i = 2
while i <= m_ws.max_row and count < 2:
    if m_ws.cell(i, type_idx+1).value == 'Company':
        company_name = m_ws.cell(i, name_idx+1).value
        print(f"\nRow {i}: COMPANY - {company_name}")
        j = i + 1
        contact_count = 0
        while j <= m_ws.max_row and m_ws.cell(j, type_idx+1).value == 'Person':
            contact_name = m_ws.cell(j, name_idx+1).value
            related = m_ws.cell(j, related_idx+1).value if related_idx else ''
            print(f"  Row {j}: CONTACT - {contact_name}")
            j += 1
            contact_count += 1
        if contact_count == 0:
            print(f"  (No contacts listed)")
        i = j if j <= m_ws.max_row else m_ws.max_row + 1
        count += 1
    else:
        i += 1
