"""Check if phones and notes are being included."""
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors_and_contacts.xlsx', data_only=True)
ws = wb.active
headers = [str(c.value or '') for c in ws[1]]

notes_idx = next((i for i, h in enumerate(headers) if 'notes' in h.lower()), None)
phone_idx = next((i for i, h in enumerate(headers) if 'phone' in h.lower()), None)
tags_idx = next((i for i, h in enumerate(headers) if 'tags' in h.lower()), None)

print("Checking merged file for phones and notes:")
print("=" * 80)

phones_count = 0
notes_count = 0
tags_count = 0

for i in range(2, ws.max_row + 1):
    company_type = ws.cell(i, 2).value
    if company_type == 'Company':
        phone = ws.cell(i, phone_idx + 1).value if phone_idx is not None else None
        notes = ws.cell(i, notes_idx + 1).value if notes_idx is not None else None
        tags = ws.cell(i, tags_idx + 1).value if tags_idx is not None else None
        
        if phone and str(phone).strip():
            phones_count += 1
        if notes and str(notes).strip():
            notes_count += 1
        if tags and str(tags).strip():
            tags_count += 1

print(f"Total companies: {ws.max_row - 1}")
print(f"Companies with phone numbers: {phones_count}")
print(f"Companies with notes: {notes_count}")
print(f"Companies with service tags: {tags_count}")

print("\nSample companies with data:")
sample_count = 0
for i in range(2, min(ws.max_row + 1, 22)):
    company_type = ws.cell(i, 2).value
    if company_type == 'Company':
        name = ws.cell(i, 1).value
        phone = ws.cell(i, phone_idx + 1).value if phone_idx is not None else None
        notes = ws.cell(i, notes_idx + 1).value if notes_idx is not None else None
        tags = ws.cell(i, tags_idx + 1).value if tags_idx is not None else None
        
        if phone or notes or tags:
            sample_count += 1
            phone_str = str(phone)[:20] if phone else "(no phone)"
            notes_str = str(notes)[:40] + "..." if notes else "(no notes)"
            tags_str = str(tags)[:30] if tags else "(no tags)"
            print(f"  {name}:")
            print(f"    Phone: {phone_str}")
            print(f"    Notes: {notes_str}")
            print(f"    Tags: {tags_str}")
            if sample_count >= 5:
                break
