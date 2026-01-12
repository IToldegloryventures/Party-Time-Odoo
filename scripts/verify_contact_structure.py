"""Verify that contacts are properly structured with their own phone/email."""
from openpyxl import load_workbook

merged = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors_and_contacts.xlsx', data_only=True)
ws = merged.active
headers = [str(c.value or '') for c in ws[1]]

name_idx = 0
type_idx = 1
related_idx = next((i for i, h in enumerate(headers) if 'related company' in h.lower()), None)
phone_idx = next((i for i, h in enumerate(headers) if 'phone' in h.lower()), None)
email_idx = next((i for i, h in enumerate(headers) if 'email' in h.lower()), None)

print("=" * 80)
print("VERIFYING CONTACT STRUCTURE")
print("=" * 80)
print("\nIn Odoo, contacts are separate records with their own phone/email fields.")
print("They are linked to companies via the 'Related Company' field.")
print("Contacts will be visible even if the parent company has no phone number.\n")

# Find companies without phones but with contacts that have phones
companies_without_phones = []
for i in range(2, ws.max_row + 1):
    if ws.cell(i, type_idx + 1).value == 'Company':
        company_name = str(ws.cell(i, name_idx + 1).value or '')
        company_phone = str(ws.cell(i, phone_idx + 1).value or '').strip() if phone_idx and ws.cell(i, phone_idx + 1).value else ''
        company_email = str(ws.cell(i, email_idx + 1).value or '').strip() if email_idx and ws.cell(i, email_idx + 1).value else ''
        
        # Check if this company has contacts with phones
        contacts_with_phones = []
        j = i + 1
        while j <= ws.max_row and ws.cell(j, type_idx + 1).value == 'Person':
            contact_name = str(ws.cell(j, name_idx + 1).value or '')
            contact_related = str(ws.cell(j, related_idx + 1).value or '').strip() if related_idx and ws.cell(j, related_idx + 1).value else ''
            contact_phone = str(ws.cell(j, phone_idx + 1).value or '').strip() if phone_idx and ws.cell(j, phone_idx + 1).value else ''
            contact_email = str(ws.cell(j, email_idx + 1).value or '').strip() if email_idx and ws.cell(j, email_idx + 1).value else ''
            
            if contact_related == company_name and (contact_phone or contact_email):
                contacts_with_phones.append({
                    'name': contact_name,
                    'phone': contact_phone,
                    'email': contact_email
                })
            j += 1
        
        if not company_phone and contacts_with_phones:
            companies_without_phones.append({
                'company': company_name,
                'company_phone': company_phone,
                'company_email': company_email,
                'contacts': contacts_with_phones
            })

print(f"Found {len(companies_without_phones)} companies WITHOUT phone numbers")
print("but WITH contacts that have phone/email information:\n")

for item in companies_without_phones[:5]:
    print(f"Company: {item['company']}")
    print(f"  Company Phone: {item['company_phone'] or '(none)'}")
    print(f"  Company Email: {item['company_email'] or '(none)'}")
    print(f"  Contacts with info:")
    for contact in item['contacts']:
        print(f"    - {contact['name']}")
        if contact['phone']:
            print(f"      Phone: {contact['phone']}")
        if contact['email']:
            print(f"      Email: {contact['email']}")
    print()

print("\n" + "=" * 80)
print("ANSWER: YES - In Odoo CRM/Sales apps, you WILL be able to see")
print("contacts and their phone numbers/emails even if the company")
print("record doesn't have a phone number. Contacts are separate records")
print("with their own contact information fields.")
print("=" * 80)
