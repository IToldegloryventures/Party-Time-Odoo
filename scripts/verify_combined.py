"""Verify the combined file format."""
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors_and_contacts.xlsx', data_only=True)
ws = wb.active
headers = [str(c.value or '') for c in ws[1]]

print("Headers:", ', '.join(headers))
print("\nFirst 10 rows (showing company then contacts):")
print("=" * 100)

for i in range(1, min(11, ws.max_row + 1)):
    name = ws.cell(i + 1, 1).value or ""
    company_type = ws.cell(i + 1, 2).value or ""
    related = ws.cell(i + 1, 3).value or ""
    reference = ws.cell(i + 1, 15).value or ""
    
    print(f"Row {i+1}: Name='{name}', Type='{company_type}', Related='{related}', Reference='{reference}'")
