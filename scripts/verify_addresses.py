"""Verify addresses are on company level, not contact level."""
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors_and_contacts.xlsx', data_only=True)
ws = wb.active
headers = [str(c.value or '') for c in ws[1]]

street_idx = next(i for i, h in enumerate(headers) if 'street' in h.lower() and '2' not in h.lower())
city_idx = next(i for i, h in enumerate(headers) if 'city' in h.lower())

print("Sample showing addresses (should be on Company, empty on Contacts):")
print("=" * 100)

for i in range(1, min(12, ws.max_row + 1)):
    name = ws.cell(i + 1, 1).value or ""
    company_type = ws.cell(i + 1, 2).value or ""
    street = ws.cell(i + 1, street_idx + 1).value or ""
    city = ws.cell(i + 1, city_idx + 1).value or ""
    
    street_display = street if street else "(empty)"
    city_display = city if city else "(empty)"
    
    print(f"Row {i+1}: Name='{name}', Type='{company_type}', Street='{street_display}', City='{city_display}'")
