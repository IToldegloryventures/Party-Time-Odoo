"""Debug why matching isn't working."""
from openpyxl import load_workbook
from difflib import SequenceMatcher

# Load all three files
rippling = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors Rippling.xlsx', data_only=True)
qbo = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors QBO.xlsx', data_only=True)
master = load_workbook(r'C:\Users\ashpt\Downloads\Master Vendors List (1).xlsx', data_only=True)

r_ws = rippling.active
q_ws = qbo.active
m_ws = master.active

# Get headers
r_headers = [str(c.value or '') for c in r_ws[1]]
q_headers = [str(c.value or '') for c in q_ws[1]]
m_headers = [str(c.value or '') for c in m_ws[1]]

# Find name columns
r_name_idx = 0
r_company_idx = 2
q_name_idx = next((i for i, h in enumerate(q_headers) if 'vendor name' in h.lower() or ('name' in h.lower() and 'company' not in h.lower())), 0)
m_name_idx = next((i for i, h in enumerate(m_headers) if 'vendor' in h.lower() and 'name' in h.lower()), 0)

# Get phone and notes columns
q_phone_idx = next((i for i, h in enumerate(q_headers) if 'phone' in h.lower()), None)
m_notes_idx = next((i for i, h in enumerate(m_headers) if 'notes' in h.lower()), None)

print("Rippling headers:", r_headers[:10])
print("QBO headers:", q_headers[:10])
print("Master headers:", m_headers[:10])
print()

# Collect Rippling vendor names
rippling_names = []
for row in r_ws.iter_rows(min_row=2):
    name_val = row[r_name_idx].value
    company_val = row[r_company_idx].value if r_company_idx < len(row) else None
    vendor_name = str(company_val or name_val or '').strip()
    if vendor_name:
        rippling_names.append(vendor_name.lower())

# Collect QBO vendor names with phones
qbo_names_phones = {}
for row in q_ws.iter_rows(min_row=2):
    name_val = row[q_name_idx].value if q_name_idx < len(row) else None
    phone_val = row[q_phone_idx].value if q_phone_idx and q_phone_idx < len(row) else None
    vendor_name = str(name_val or '').strip()
    if vendor_name:
        qbo_names_phones[vendor_name.lower()] = str(phone_val or '').strip()

# Collect Master vendor names with notes
master_names_notes = {}
for row in m_ws.iter_rows(min_row=2):
    name_val = row[m_name_idx].value if m_name_idx < len(row) else None
    notes_val = row[m_notes_idx].value if m_notes_idx and m_notes_idx < len(row) else None
    vendor_name = str(name_val or '').strip()
    if vendor_name:
        master_names_notes[vendor_name.lower()] = str(notes_val or '').strip()

print(f"Rippling vendors: {len(rippling_names)}")
print(f"QBO vendors with phones: {sum(1 for p in qbo_names_phones.values() if p)}")
print(f"Master vendors with notes: {sum(1 for n in master_names_notes.values() if n)}")
print()

# Try exact matching
exact_qbo_matches = sum(1 for r_name in rippling_names if r_name in qbo_names_phones)
exact_master_matches = sum(1 for r_name in rippling_names if r_name in master_names_notes)

print(f"Exact matches - QBO: {exact_qbo_matches}, Master: {exact_master_matches}")
print()

# Try fuzzy matching (similarity > 0.85)
def normalize_name(name):
    """Normalize name for matching."""
    name = name.lower().strip()
    # Remove common suffixes
    import re
    name = re.sub(r'\s+(inc|llc|ltd|corp|corporation|company|co)\.?$', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

fuzzy_qbo_matches = 0
fuzzy_master_matches = 0

for r_name in rippling_names[:20]:  # Check first 20
    r_norm = normalize_name(r_name)
    best_qbo_match = None
    best_qbo_score = 0
    best_master_match = None
    best_master_score = 0
    
    for q_name in qbo_names_phones.keys():
        q_norm = normalize_name(q_name)
        score = SequenceMatcher(None, r_norm, q_norm).ratio()
        if score > best_qbo_score:
            best_qbo_score = score
            best_qbo_match = q_name
    
    for m_name in master_names_notes.keys():
        m_norm = normalize_name(m_name)
        score = SequenceMatcher(None, r_norm, m_norm).ratio()
        if score > best_master_score:
            best_master_score = score
            best_master_match = m_name
    
    if best_qbo_score > 0.85:
        fuzzy_qbo_matches += 1
        print(f"QBO match: '{r_name}' -> '{best_qbo_match}' (score: {best_qbo_score:.2f})")
    if best_master_score > 0.85:
        fuzzy_master_matches += 1
        print(f"Master match: '{r_name}' -> '{best_master_match}' (score: {best_master_score:.2f})")

print(f"\nFuzzy matches (first 20 Rippling): QBO: {fuzzy_qbo_matches}, Master: {fuzzy_master_matches}")
