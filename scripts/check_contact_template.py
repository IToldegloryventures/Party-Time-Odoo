"""Check Odoo contact import template structure."""
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\ashpt\Downloads\contacts_import_template.xlsx', data_only=True)
ws = wb.active
headers = [str(c.value or '') for c in ws[1]]

print("Odoo Contact Import Template - All columns:")
for i, h in enumerate(headers):
    if h:
        print(f"  {i+1}. {h}")

print("\nChecking for ID/external ID fields...")
id_fields = [h for h in headers if h and ('id' in h.lower() or 'external' in h.lower() or 'parent' in h.lower())]
print(f"ID/Parent fields found: {id_fields if id_fields else 'None'}")

print("\n'Related Company' field details:")
related_company_idx = next((i for i, h in enumerate(headers) if h and 'related company' in h.lower()), None)
if related_company_idx is not None:
    print(f"  Found at column {related_company_idx + 1}: '{headers[related_company_idx]}'")
    print("  This field links contacts to parent companies by NAME")
    print("  Odoo will search for a company with this name to link the contact")
else:
    print("  NOT FOUND - this is a problem!")
