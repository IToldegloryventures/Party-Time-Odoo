"""Quick script to auto-approve all fuzzy matches and complete the merge."""
from openpyxl import load_workbook
import sys

review_file = r"C:\Users\ashpt\Downloads\vendor_merge_output\fuzzy_matches_review.xlsx"

try:
    wb = load_workbook(review_file)
    ws = wb.active
    
    # Mark all fuzzy matches as "Merge"
    for row in ws.iter_rows(min_row=2, values_only=False):
        decision_cell = row[1]  # Decision column
        if not decision_cell.value or str(decision_cell.value).strip() == "":
            decision_cell.value = "Merge"
    
    wb.save(review_file)
    print(f"[OK] Auto-approved all fuzzy matches in {review_file}")
    print("Now run the merge script again to complete the process.")
    
except Exception as e:
    print(f"[ERROR] Could not auto-approve: {e}")
    sys.exit(1)
