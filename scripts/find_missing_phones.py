"""Find exactly which QBO phones are missing from merged file."""
from openpyxl import load_workbook
from difflib import SequenceMatcher
import re

def normalize_name(name):
    """Normalize name for matching."""
    if not name:
        return ""
    name = str(name).strip().lower()
    name = re.sub(r'\s+(inc|llc|ltd|corp|corporation|company|co|pllc)\.?$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'[^\w\s]', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

# Load QBO
qbo = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors QBO.xlsx', data_only=True)
q_ws = qbo.active
q_headers = [str(c.value or '') for c in q_ws[1]]
q_name_idx = next((i for i, h in enumerate(q_headers) if 'vendor name' in h.lower()), 0)
q_phone_idx = next((i for i, h in enumerate(q_headers) if 'phone' in h.lower()), None)

# Load Rippling
rippling = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors Rippling.xlsx', data_only=True)
r_ws = rippling.active
r_name_idx = 0
r_company_idx = 2

# Load merged
merged = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors_and_contacts.xlsx', data_only=True)
m_ws = merged.active
m_headers = [str(c.value or '') for c in m_ws[1]]
m_name_idx = 0
m_phone_idx = next((i for i, h in enumerate(m_headers) if 'phone' in h.lower()), None)

# Build Rippling vendor name map (normalized -> original)
rippling_map = {}
for row in r_ws.iter_rows(min_row=2):
    name_val = row[r_name_idx].value
    company_val = row[r_company_idx].value if r_company_idx < len(row) else None
    vendor_name = str(company_val or name_val or '').strip()
    if vendor_name:
        rippling_map[normalize_name(vendor_name)] = vendor_name

# Build merged vendor phone map (normalized name -> phone)
merged_phones = {}
for i in range(2, m_ws.max_row + 1):
    if m_ws.cell(i, 2).value == 'Company':
        name = str(m_ws.cell(i, m_name_idx + 1).value or '').strip()
        phone = str(m_ws.cell(i, m_phone_idx + 1).value or '').strip() if m_phone_idx and m_ws.cell(i, m_phone_idx + 1).value else ''
        if name:
            merged_phones[normalize_name(name)] = phone

# Find QBO vendors with phones that should match Rippling but don't have phones in merged
missing_phones = []
qbo_phones_no_match = []

for row in q_ws.iter_rows(min_row=2):
    q_name = str(row[q_name_idx].value or '').strip() if q_name_idx < len(row) else ''
    q_phone = str(row[q_phone_idx].value or '').strip() if q_phone_idx and q_phone_idx < len(row) and row[q_phone_idx].value else ''
    
    if not q_name or not q_phone:
        continue
    
    q_norm = normalize_name(q_name)
    
    # Check if this QBO vendor matches any Rippling vendor
    best_rippling_match = None
    best_score = 0.0
    for r_norm, r_name in rippling_map.items():
        score = SequenceMatcher(None, q_norm, r_norm).ratio()
        if score > best_score:
            best_score = score
            best_rippling_match = (r_norm, r_name)
    
    if best_score > 0.80:  # Should match
        # Check if phone is in merged file
        merged_phone = merged_phones.get(best_rippling_match[0], '')
        if not merged_phone or merged_phone != q_phone:
            missing_phones.append({
                'qbo_name': q_name,
                'qbo_phone': q_phone,
                'rippling_name': best_rippling_match[1],
                'merged_phone': merged_phone,
                'score': best_score
            })
    else:
        # This QBO vendor doesn't match any Rippling vendor (expected - we only use Rippling as base)
        qbo_phones_no_match.append((q_name, q_phone))

print("=" * 80)
print("MISSING PHONES ANALYSIS")
print("=" * 80)
print(f"\nQBO vendors with phones that match Rippling but phone is missing/incorrect in merged:")
print(f"  Found {len(missing_phones)} cases")
for item in missing_phones[:10]:
    print(f"\n  QBO: '{item['qbo_name']}' -> Phone: '{item['qbo_phone']}'")
    print(f"  Matches Rippling: '{item['rippling_name']}' (score: {item['score']:.2f})")
    merged_phone_display = item['merged_phone'] or "(none)"
    print(f"  Merged phone: '{merged_phone_display}'")

print(f"\n\nQBO vendors with phones that DON'T match any Rippling vendor (can't be added):")
print(f"  Found {len(qbo_phones_no_match)} cases")
for q_name, q_phone in qbo_phones_no_match[:5]:
    print(f"    '{q_name}' -> Phone: '{q_phone}'")

print(f"\n\nSUMMARY:")
print(f"  QBO vendors with phones: 85")
print(f"  Phones that should be in merged but aren't: {len(missing_phones)}")
print(f"  Phones from vendors not in Rippling (can't add): {len(qbo_phones_no_match)}")
