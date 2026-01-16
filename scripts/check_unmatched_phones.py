"""Check for unmatched phones and emails from QBO."""
from openpyxl import load_workbook
from difflib import SequenceMatcher
import re

def normalize_name(name):
    """Normalize name for matching."""
    if not name:
        return ""
    name = str(name).strip().lower()
    name = re.sub(r'\s+(inc|llc|ltd|corp|corporation|company|co)\.?$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'[^\w\s]', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

# Load QBO
qbo = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors QBO.xlsx', data_only=True)
q_ws = qbo.active
q_headers = [str(c.value or '') for c in q_ws[1]]
q_name_idx = next((i for i, h in enumerate(q_headers) if 'vendor name' in h.lower()), 0)
q_phone_idx = next((i for i, h in enumerate(q_headers) if 'phone' in h.lower()), None)
q_email_idx = next((i for i, h in enumerate(q_headers) if 'email' in h.lower()), None)

# Load Rippling
rippling = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors Rippling.xlsx', data_only=True)
r_ws = rippling.active
r_headers = [str(c.value or '') for c in r_ws[1]]
r_name_idx = 0
r_company_idx = 2

# Load merged
merged = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors_and_contacts.xlsx', data_only=True)
m_ws = merged.active
m_headers = [str(c.value or '') for c in m_ws[1]]
m_name_idx = 0
m_phone_idx = next((i for i, h in enumerate(m_headers) if 'phone' in h.lower()), None)
m_email_idx = next((i for i, h in enumerate(m_headers) if 'email' in h.lower()), None)

# Get all Rippling vendor names (normalized)
rippling_names = {}
for row in r_ws.iter_rows(min_row=2):
    name_val = row[r_name_idx].value
    company_val = row[r_company_idx].value if r_company_idx < len(row) else None
    vendor_name = str(company_val or name_val or '').strip()
    if vendor_name:
        rippling_names[normalize_name(vendor_name)] = vendor_name

# Get all merged vendor names (normalized)
merged_names = {}
for i in range(2, m_ws.max_row + 1):
    if m_ws.cell(i, 2).value == 'Company':
        name = str(m_ws.cell(i, m_name_idx + 1).value or '').strip()
        if name:
            merged_names[normalize_name(name)] = name

# Check QBO vendors with phones/emails that weren't matched
unmatched_phones = []
unmatched_emails = []

for row in q_ws.iter_rows(min_row=2):
    q_name = str(row[q_name_idx].value or '').strip() if q_name_idx < len(row) else ''
    q_phone = str(row[q_phone_idx].value or '').strip() if q_phone_idx and q_phone_idx < len(row) and row[q_phone_idx].value else ''
    q_email = str(row[q_email_idx].value or '').strip() if q_email_idx and q_email_idx < len(row) and row[q_email_idx].value else ''
    
    if not q_name:
        continue
    
    q_norm = normalize_name(q_name)
    
    # Check if this QBO vendor matches any Rippling vendor
    matched = False
    for r_norm, r_name in rippling_names.items():
        if q_norm == r_norm or SequenceMatcher(None, q_norm, r_norm).ratio() > 0.80:
            matched = True
            break
    
    if matched:
        # Check if phone/email made it to merged file
        found_in_merged = False
        for m_norm, m_name in merged_names.items():
            if SequenceMatcher(None, q_norm, m_norm).ratio() > 0.80:
                # Check if this merged vendor has the phone/email
                for i in range(2, m_ws.max_row + 1):
                    if m_ws.cell(i, 2).value == 'Company':
                        m_name_check = str(m_ws.cell(i, m_name_idx + 1).value or '').strip()
                        if normalize_name(m_name_check) == m_norm:
                            m_phone = str(m_ws.cell(i, m_phone_idx + 1).value or '').strip() if m_phone_idx else ''
                            m_email = str(m_ws.cell(i, m_email_idx + 1).value or '').strip() if m_email_idx else ''
                            
                            if q_phone and m_phone != q_phone:
                                unmatched_phones.append((q_name, q_phone, m_name_check, m_phone))
                            if q_email and m_email != q_email:
                                unmatched_emails.append((q_name, q_email, m_name_check, m_email))
                            found_in_merged = True
                            break
                if found_in_merged:
                    break
    else:
        # This QBO vendor doesn't match any Rippling vendor (expected - we only use Rippling as base)
        pass

print("=" * 80)
print("UNMATCHED PHONES AND EMAILS ANALYSIS")
print("=" * 80)
print(f"\nQBO vendors with phones that might not be in merged file:")
print(f"  Found {len(unmatched_phones)} potential missing phones")
if unmatched_phones:
    for q_name, q_phone, m_name, m_phone in unmatched_phones[:10]:
        print(f"    QBO: '{q_name}' -> Phone: '{q_phone}'")
        m_phone_display = m_phone or "(none)"
        print(f"    Merged: '{m_name}' -> Phone: '{m_phone_display}'")

print(f"\nQBO vendors with emails that might not be in merged file:")
print(f"  Found {len(unmatched_emails)} potential missing emails")
if unmatched_emails:
    for q_name, q_email, m_name, m_email in unmatched_emails[:10]:
        print(f"    QBO: '{q_name}' -> Email: '{q_email}'")
        m_email_display = m_email or "(none)"
        print(f"    Merged: '{m_name}' -> Email: '{m_email_display}'")

# Also check total counts
qbo_phones_count = sum(1 for row in q_ws.iter_rows(min_row=2) if q_phone_idx and row[q_phone_idx].value and str(row[q_phone_idx].value).strip())
qbo_emails_count = sum(1 for row in q_ws.iter_rows(min_row=2) if q_email_idx and row[q_email_idx].value and str(row[q_email_idx].value).strip())

merged_phones_count = sum(1 for i in range(2, m_ws.max_row + 1) if m_ws.cell(i, 2).value == 'Company' and m_phone_idx and m_ws.cell(i, m_phone_idx + 1).value and str(m_ws.cell(i, m_phone_idx + 1).value).strip())
merged_emails_count = sum(1 for i in range(2, m_ws.max_row + 1) if m_ws.cell(i, 2).value == 'Company' and m_email_idx and m_ws.cell(i, m_email_idx + 1).value and str(m_ws.cell(i, m_email_idx + 1).value).strip())

print(f"\nSUMMARY:")
print(f"  QBO file has {qbo_phones_count} vendors with phones")
print(f"  Merged file has {merged_phones_count} companies with phones")
print(f"  Missing: {qbo_phones_count - merged_phones_count} phones")
print(f"\n  QBO file has {qbo_emails_count} vendors with emails")
print(f"  Merged file has {merged_emails_count} companies with emails")
print(f"  Missing: {qbo_emails_count - merged_emails_count} emails")
