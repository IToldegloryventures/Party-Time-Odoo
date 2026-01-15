"""Check if vendor notes from Master Vendor List are preserved correctly."""
from openpyxl import load_workbook

# Load Master Vendor List
master_wb = load_workbook(r'C:\Users\ashpt\Downloads\Master Vendors List (1).xlsx', data_only=True)
master_ws = master_wb.active
master_headers = [cell.value for cell in master_ws[1]]

# Find Notes column
notes_col_idx = None
for i, header in enumerate(master_headers):
    if header and 'notes' in str(header).lower():
        notes_col_idx = i
        break

# Find Organization column
org_col_idx = None
for i, header in enumerate(master_headers):
    if header and ('organization' in str(header).lower() or 'vendor' in str(header).lower()):
        org_col_idx = i
        break

print(f"Master Vendor List - Notes column: {notes_col_idx}, Org column: {org_col_idx}")

# Get master vendor notes
master_notes = {}
for row in master_ws.iter_rows(min_row=2, values_only=False):
    org = row[org_col_idx].value if org_col_idx is not None else None
    notes = row[notes_col_idx].value if notes_col_idx is not None else None
    if org and notes:
        master_notes[str(org).strip()] = str(notes).strip()

print(f"\nFound {len(master_notes)} vendors with notes in Master Vendor List")
print("\nSample notes from Master Vendor List:")
for i, (org, notes) in enumerate(list(master_notes.items())[:5]):
    print(f"  {org}: {notes[:80]}...")

# Load merged vendors
merged_wb = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors.xlsx', data_only=True)
merged_ws = merged_wb.active
merged_headers = [cell.value for cell in merged_ws[1]]

# Find notes column in merged file
merged_notes_col_idx = None
for i, header in enumerate(merged_headers):
    if header and ('notes' in str(header).lower() or 'x_vendor_notes' in str(header).lower()):
        merged_notes_col_idx = i
        break

# Find name column
merged_name_col_idx = None
for i, header in enumerate(merged_headers):
    if header and 'name' in str(header).lower() and 'company' not in str(header).lower():
        merged_name_col_idx = i
        break

print(f"\nMerged file - Notes column: {merged_notes_col_idx}, Name column: {merged_name_col_idx}")

# Compare notes
matches = 0
mismatches = []
for row in merged_ws.iter_rows(min_row=2, values_only=False):
    vendor_name = str(row[merged_name_col_idx].value or "").strip()
    merged_notes = str(row[merged_notes_col_idx].value or "").strip()
    
    if vendor_name in master_notes:
        master_note = master_notes[vendor_name]
        if merged_notes == master_note:
            matches += 1
        else:
            mismatches.append((vendor_name, master_note[:50], merged_notes[:50]))

print(f"\nComparison Results:")
print(f"  Vendors with notes in Master List: {len(master_notes)}")
print(f"  Exact matches in merged file: {matches}")
print(f"  Mismatches: {len(mismatches)}")

if mismatches:
    print("\nMismatches found:")
    for vendor, master_note, merged_note in mismatches[:10]:
        print(f"  {vendor}:")
        print(f"    Master: {master_note}...")
        print(f"    Merged: {merged_note}...")
