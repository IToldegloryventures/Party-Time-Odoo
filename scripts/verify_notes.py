"""Verify that Master Vendor List notes are in the merged file."""
from openpyxl import load_workbook

# Check merged file
wb = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors.xlsx', data_only=True)
ws = wb.active
headers = [str(c.value or '') for c in ws[1]]

print("Columns in merged_vendors.xlsx:")
for i, h in enumerate(headers):
    if h:
        print(f"  {i+1}. {h}")

# Find notes column
notes_idx = None
for i, h in enumerate(headers):
    if h and ('x_vendor_notes' in h.lower() or ('notes' in h.lower() and 'vendor' in h.lower())):
        notes_idx = i
        break

if notes_idx is None:
    print("\n[ERROR] x_vendor_notes column NOT FOUND in merged file!")
else:
    print(f"\n[OK] Found x_vendor_notes column at index {notes_idx + 1}")
    
    # Count vendors with notes
    vendors_with_notes = 0
    print("\nSample vendors with notes from merged file:")
    for i in range(2, min(ws.max_row + 1, 12)):  # First 10 rows
        vendor_name = ws.cell(i, 1).value
        notes = ws.cell(i, notes_idx + 1).value
        if notes:
            vendors_with_notes += 1
            notes_str = str(notes)[:80]
            print(f"  {vendor_name}: {notes_str}...")
    
    print(f"\nTotal vendors with notes in merged file: {vendors_with_notes} (out of {ws.max_row - 1} vendors)")
