"""Verify the merge used correct priority: Rippling base, QBO phones, Master notes."""
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors.xlsx', data_only=True)
ws = wb.active
headers = [str(c.value or '') for c in ws[1]]

name_idx = 0
phone_idx = next(i for i, h in enumerate(headers) if 'phone' in h.lower())
notes_idx = next((i for i, h in enumerate(headers) if 'x_vendor_notes' in h.lower()), None)
service_idx = next((i for i, h in enumerate(headers) if 'x_vendor_service_tag_ids' in h.lower()), None)

print("Sample merged vendors (should show Rippling as base, QBO phones, Master notes/service tags):")
print("=" * 80)

for i in range(2, min(ws.max_row + 1, 12)):  # First 10 vendors
    name = ws.cell(i, name_idx + 1).value
    phone = ws.cell(i, phone_idx + 1).value or "(no phone)"
    notes = ws.cell(i, notes_idx + 1).value[:60] + "..." if notes_idx and ws.cell(i, notes_idx + 1).value else "(no notes)"
    services = ws.cell(i, service_idx + 1).value or "(no services)"
    
    print(f"\n{name}:")
    print(f"  Phone: {phone}")
    print(f"  Notes: {notes}")
    print(f"  Services: {services}")
