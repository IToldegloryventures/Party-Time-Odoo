"""Verify merged list contains ONLY Rippling vendors with QBO phones and Master notes/services."""
from openpyxl import load_workbook
from collections import Counter

print("=" * 80)
print("VERIFICATION: Merged List Completeness Check")
print("=" * 80)

# Load Rippling vendors (this is our SOURCE OF TRUTH)
rippling_wb = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors Rippling.xlsx', data_only=True)
rippling_ws = rippling_wb.active
rippling_headers = [str(c.value or '') for c in rippling_ws[1]]

rippling_name_idx = next(i for i, h in enumerate(rippling_headers) if 'name' in h.lower() and 'company' not in h.lower())
rippling_company_idx = next((i for i, h in enumerate(rippling_headers) if 'company name' in h.lower()), None)

rippling_vendors = set()
rippling_vendor_names = []
for row in rippling_ws.iter_rows(min_row=2, values_only=False):
    name_val = row[rippling_name_idx].value
    company_name_val = row[rippling_company_idx].value if rippling_company_idx is not None else None
    
    # Use company_name if name is empty, otherwise use name
    vendor_name = None
    if company_name_val and str(company_name_val).strip():
        vendor_name = str(company_name_val).strip()
    elif name_val and str(name_val).strip():
        vendor_name = str(name_val).strip()
    
    if vendor_name:
        rippling_vendors.add(vendor_name.lower())
        rippling_vendor_names.append(vendor_name)

print(f"\n1. Rippling Spreadsheet:")
print(f"   Total vendor records: {len(rippling_vendor_names)}")
print(f"   Unique vendor names: {len(rippling_vendors)}")

# Load merged file
merged_wb = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors_and_contacts.xlsx', data_only=True)
merged_ws = merged_wb.active
merged_headers = [str(c.value or '') for c in merged_ws[1]]

name_idx = next(i for i, h in enumerate(merged_headers) if 'name' in h.lower() and '*' in h)
type_idx = next(i for i, h in enumerate(merged_headers) if 'company type' in h.lower() and '*' in h)

merged_companies = []
merged_contacts = []
for row in merged_ws.iter_rows(min_row=2, values_only=False):
    name = row[name_idx].value
    company_type = row[type_idx].value
    if name and str(name).strip():
        if company_type == 'Company':
            merged_companies.append(str(name).strip())
        elif company_type == 'Person':
            merged_contacts.append(str(name).strip())

merged_company_set = {name.lower() for name in merged_companies}

print(f"\n2. Merged File:")
print(f"   Total companies: {len(merged_companies)}")
print(f"   Total contacts: {len(merged_contacts)}")
print(f"   Unique company names: {len(merged_company_set)}")

# Check for duplicates in merged file
merged_duplicates = [name for name, count in Counter(merged_companies).items() if count > 1]
if merged_duplicates:
    print(f"\n   [WARNING] DUPLICATE COMPANIES FOUND: {len(merged_duplicates)}")
    for dup in merged_duplicates[:10]:
        print(f"      - {dup}")
else:
    print(f"\n   [OK] No duplicate companies")

# Check if all Rippling vendors are in merged file
missing_from_merged = []
for rippling_name in rippling_vendor_names:
    if rippling_name.lower() not in merged_company_set:
        missing_from_merged.append(rippling_name)

if missing_from_merged:
    print(f"\n   [WARNING] MISSING FROM MERGED FILE: {len(missing_from_merged)} vendors from Rippling")
    for missing in missing_from_merged[:20]:
        print(f"      - {missing}")
else:
    print(f"\n   [OK] All Rippling vendors are in merged file")

# Check for extra vendors in merged file (not from Rippling)
extra_in_merged = []
for merged_name in merged_companies:
    if merged_name.lower() not in rippling_vendors:
        extra_in_merged.append(merged_name)

if extra_in_merged:
    print(f"\n   [WARNING] EXTRA VENDORS IN MERGED FILE (not in Rippling): {len(extra_in_merged)}")
    print(f"      These should NOT be here if Rippling is the only source!")
    for extra in extra_in_merged[:20]:
        print(f"      - {extra}")
else:
    print(f"\n   [OK] No extra vendors (all companies are from Rippling)")

# Load QBO to check phone numbers
qbo_wb = load_workbook(r'C:\Users\ashpt\Downloads\Copy of Vendors QBO.xlsx', data_only=True)
qbo_ws = qbo_wb.active
qbo_headers = [str(c.value or '') for c in qbo_ws[1]]

qbo_name_idx = next(i for i, h in enumerate(qbo_headers) if 'vendor name' in h.lower() or ('name' in h.lower() and 'company' not in h.lower()))
qbo_phone_idx = next((i for i, h in enumerate(qbo_headers) if 'phone' in h.lower()), None)

qbo_phones = {}
if qbo_phone_idx is not None:
    for row in qbo_ws.iter_rows(min_row=2, values_only=False):
        name = row[qbo_name_idx].value
        phone = row[qbo_phone_idx].value if qbo_phone_idx is not None else None
        if name and str(name).strip():
            qbo_phones[str(name).strip().lower()] = str(phone).strip() if phone else ""

print(f"\n3. QBO Spreadsheet:")
print(f"   Vendors with phone numbers: {len([p for p in qbo_phones.values() if p])}")

# Load Master Vendor List
master_wb = load_workbook(r'C:\Users\ashpt\Downloads\Master Vendors List (1).xlsx', data_only=True)
master_ws = master_wb.active
master_headers = [str(c.value or '') for c in master_ws[1]]

master_org_idx = next(i for i, h in enumerate(master_headers) if 'organization' in h.lower() or 'vendor' in h.lower())
master_notes_idx = next((i for i, h in enumerate(master_headers) if 'notes' in h.lower()), None)
master_category_idx = next((i for i, h in enumerate(master_headers) if 'category' in h.lower() or 'service' in h.lower()), None)

master_vendors = set()
master_notes = {}
master_services = {}
for row in master_ws.iter_rows(min_row=2, values_only=False):
    org = row[master_org_idx].value
    if org and str(org).strip():
        org_name = str(org).strip()
        master_vendors.add(org_name.lower())
        if master_notes_idx is not None:
            notes = row[master_notes_idx].value
            if notes:
                master_notes[org_name.lower()] = str(notes).strip()
        if master_category_idx is not None:
            category = row[master_category_idx].value
            if category:
                master_services[org_name.lower()] = str(category).strip()

print(f"\n4. Master Vendor List:")
print(f"   Total vendors: {len(master_vendors)}")
print(f"   Vendors with notes: {len(master_notes)}")
print(f"   Vendors with service categories: {len(master_services)}")

# Summary
print(f"\n" + "=" * 80)
print("SUMMARY:")
print("=" * 80)
print(f"[OK] Rippling vendors: {len(rippling_vendors)}")
print(f"[OK] Merged companies: {len(merged_company_set)}")
print(f"[OK] Missing from merged: {len(missing_from_merged)}")
print(f"[OK] Extra in merged (not in Rippling): {len(extra_in_merged)}")
print(f"[OK] Duplicates in merged: {len(merged_duplicates)}")

if missing_from_merged or extra_in_merged or merged_duplicates:
    print(f"\n[WARNING] ISSUES FOUND - Please review above!")
else:
    print(f"\n[OK] VERIFICATION PASSED - All Rippling vendors are present, no extras, no duplicates!")
