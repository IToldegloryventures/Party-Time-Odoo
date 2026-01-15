"""Verify that contacts/talent are in the merged vendors file."""
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\ashpt\Downloads\vendor_merge_output\merged_vendors.xlsx', data_only=True)
ws = wb.active
headers = [str(c.value or '') for c in ws[1]]

print("Columns in merged_vendors.xlsx:")
for i, h in enumerate(headers):
    if h:
        print(f"  {i+1}. {h}")

# Find contacts column
contacts_idx = None
for i, h in enumerate(headers):
    if h and ('contact' in h.lower() or 'talent' in h.lower() or 'x_vendor_contacts' in h.lower()):
        contacts_idx = i
        break

if contacts_idx is None:
    print("\n[ERROR] Contacts/Talent column NOT FOUND!")
else:
    print(f"\n[OK] Found contacts column at index {contacts_idx + 1}: '{headers[contacts_idx]}'")
    
    # Show sample vendors with contacts
    print("\nSample vendors with contacts/talent:")
    count = 0
    for i in range(2, min(ws.max_row + 1, 22)):  # First 20 rows
        vendor_name = ws.cell(i, 1).value
        contacts = ws.cell(i, contacts_idx + 1).value
        if contacts:
            count += 1
            contacts_str = str(contacts)[:100]
            print(f"  {vendor_name}: {contacts_str}")
    
    print(f"\nTotal vendors with contacts in first 20 rows: {count}")
