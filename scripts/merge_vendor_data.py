#!/usr/bin/env python3
"""
Vendor Data Merge Script

Merges vendor data from multiple Excel sources (QBO, Rippling, Master Vendor List)
using exact matching first, then fuzzy matching with manual review.

Usage:
    python scripts/merge_vendor_data.py [--qbo PATH] [--rippling PATH] [--master PATH] [--output-dir PATH]
"""

import re
import sys
import json
import difflib
import argparse
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment


@dataclass
class VendorRecord:
    """Represents a vendor record from a source file."""
    source: str  # 'qbo', 'rippling', 'master'
    name: str
    is_company: bool = True
    company_name: str = ""
    country: str = ""
    state: str = ""
    city: str = ""
    street: str = ""
    street2: str = ""
    zip: str = ""
    phone: str = ""
    email: str = ""
    service_tags: List[str] = field(default_factory=list)
    vendor_notes: str = ""
    contacts: List[Dict] = field(default_factory=list)
    row_index: int = 0
    
    def normalize_name(self) -> str:
        """Normalize vendor name for matching."""
        name = self.name.strip().lower()
        # Remove common suffixes/prefixes
        name = re.sub(r'\s+(inc|llc|ltd|corp|corporation|company|co)\.?$', '', name, flags=re.IGNORECASE)
        # Remove extra whitespace
        name = re.sub(r'\s+', ' ', name)
        return name.strip()
    
    def normalize_phone(self) -> str:
        """Normalize phone number (digits only)."""
        return re.sub(r'\D', '', self.phone) if self.phone else ""
    
    def normalize_email(self) -> str:
        """Normalize email (lowercase, trimmed)."""
        return self.email.strip().lower() if self.email else ""


class VendorMerger:
    """Handles merging vendor data from multiple sources."""
    
    def __init__(self, qbo_path: str, rippling_path: str, master_path: str, 
                 vendor_template_path: str, contact_template_path: str,
                 output_dir: str = "."):
        self.qbo_path = Path(qbo_path)
        self.rippling_path = Path(rippling_path)
        self.master_path = Path(master_path)
        self.vendor_template_path = Path(vendor_template_path)
        self.contact_template_path = Path(contact_template_path)
        self.output_dir = Path(output_dir)
        
        self.vendors: List[VendorRecord] = []
        self.exact_matches: Dict[str, List[VendorRecord]] = defaultdict(list)
        self.fuzzy_matches: List[Tuple[VendorRecord, VendorRecord, float]] = []
        self.merged_vendors: List[VendorRecord] = []
        
    def load_qbo_vendors(self) -> List[VendorRecord]:
        """Load vendors from QBO spreadsheet."""
        wb = load_workbook(self.qbo_path, data_only=True)
        ws = wb.active
        
        # Read header row
        headers = [cell.value for cell in ws[1]]
        header_map = {}
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                # Check for vendor name first (more specific)
                if 'vendor name' in header_lower:
                    header_map['name'] = i
                elif 'company name' in header_lower:
                    header_map['company_name'] = i
                elif 'phone' in header_lower:
                    header_map['phone'] = i
                elif 'email' in header_lower:
                    header_map['email'] = i
                elif 'country' in header_lower:
                    header_map['country'] = i
                elif 'state' in header_lower:
                    header_map['state'] = i
                elif 'city' in header_lower:
                    header_map['city'] = i
                elif 'street' in header_lower or ('address' in header_lower and '2' not in header_lower):
                    header_map['street'] = i
                elif 'zip' in header_lower:
                    header_map['zip'] = i
                elif 'is_company' in header_lower:
                    header_map['is_company'] = i
                # Fallback: if no vendor name found yet and this is just "name" (not company name)
                elif 'name' in header_lower and 'company' not in header_lower and 'name' not in header_map:
                    header_map['name'] = i
        
        vendors = []
        # Debug: Check header map
        if 'name' not in header_map:
            print(f"[WARNING] QBO: Could not find name column. Headers: {[str(h) for h in headers[:15]]}")
            print(f"[WARNING] QBO: Header map: {header_map}")
            return vendors  # Can't proceed without name column
        
        name_col_idx = header_map.get('name')
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if name_col_idx is None or name_col_idx >= len(row):
                continue
            name = row[name_col_idx].value
            if not name or not str(name).strip():
                continue
            
            vendor = VendorRecord(
                source='qbo',
                name=str(name).strip(),
                is_company=row[header_map.get('is_company', 1)].value if header_map.get('is_company') is not None else True,
                company_name=str(row[header_map.get('company_name', 2)].value or "").strip(),
                country=str(row[header_map.get('country', 3)].value or "").strip(),
                state=str(row[header_map.get('state', 4)].value or "").strip(),
                city=str(row[header_map.get('city', 5)].value or "").strip(),
                street=str(row[header_map.get('street', 6)].value or "").strip(),
                zip=str(row[header_map.get('zip', 7)].value or "").strip(),
                phone=str(row[header_map.get('phone', 8)].value or "").strip(),
                email=str(row[header_map.get('email', 9)].value or "").strip(),
                row_index=row_idx
            )
            vendors.append(vendor)
        
        return vendors
    
    def load_rippling_vendors(self) -> List[VendorRecord]:
        """Load vendors from Rippling spreadsheet."""
        wb = load_workbook(self.rippling_path, data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        header_map = {}
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                if 'name' in header_lower and 'company' not in header_lower:
                    header_map['name'] = i
                elif 'company name' in header_lower:
                    header_map['company_name'] = i
                elif 'contact email' in header_lower or 'email' in header_lower:
                    header_map['email'] = i
                elif 'country' in header_lower:
                    header_map['country'] = i
                elif 'state' in header_lower:
                    header_map['state'] = i
                elif 'city' in header_lower:
                    header_map['city'] = i
                elif 'address' in header_lower and '2' not in header_lower:
                    header_map['street'] = i
                elif 'address 2' in header_lower or 'address2' in header_lower:
                    header_map['street2'] = i
                elif 'zip' in header_lower:
                    header_map['zip'] = i
                elif 'is_company' in header_lower:
                    header_map['is_company'] = i
        
        vendors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            name_val = row[header_map.get('name', 0)].value if header_map.get('name') is not None else None
            company_name_val = row[header_map.get('company_name', 2)].value if header_map.get('company_name') is not None else None
            is_company_val = row[header_map.get('is_company', 1)].value if header_map.get('is_company') is not None else None
            
            # Determine the vendor name - prefer company_name, then name
            vendor_name = None
            if company_name_val and str(company_name_val).strip():
                vendor_name = str(company_name_val).strip()
            elif name_val and str(name_val).strip():
                vendor_name = str(name_val).strip()
            
            # Skip rows with no name at all
            if not vendor_name:
                continue
            
            # Determine if this is a company or contact
            is_company = bool(is_company_val) if is_company_val is not None else True
            
            # Use vendor_name for both name and company_name fields
            name = vendor_name
            company_name_str = company_name_val if company_name_val and str(company_name_val).strip() else vendor_name
            
            # Extract contact information if this is a contact (not a company)
            contacts = []
            if not is_company and name:
                # This is a contact person, not a company
                contact_email = str(row[header_map.get('email', 9)].value or "").strip() if header_map.get('email') is not None else ""
                contacts.append({
                    'name': name,
                    'email': contact_email,
                    'phone': "",  # Rippling doesn't have phone for contacts
                })
                # For contacts, the company_name is the parent company
                # We'll create a vendor record for the company if it doesn't exist
                # But for now, store the contact with the company name
                if company_name_str:
                    name = company_name_str  # Use company name as vendor name
                else:
                    name = name  # Use contact name if no company
            
            vendor = VendorRecord(
                source='rippling',
                name=name,
                is_company=True,  # Always treat as company for vendor records
                company_name=company_name_str if company_name_str else name,
                country=str(row[header_map.get('country', 3)].value or "").strip() if header_map.get('country') is not None else "",
                state=str(row[header_map.get('state', 4)].value or "").strip() if header_map.get('state') is not None else "",
                city=str(row[header_map.get('city', 5)].value or "").strip() if header_map.get('city') is not None else "",
                street=str(row[header_map.get('street', 7)].value or "").strip() if header_map.get('street') is not None else "",
                street2=str(row[header_map.get('street2', 8)].value or "").strip() if header_map.get('street2') is not None else "",
                zip=str(row[header_map.get('zip', 6)].value or "").strip() if header_map.get('zip') is not None else "",
                email=str(row[header_map.get('email', 9)].value or "").strip() if header_map.get('email') is not None else "",
                contacts=contacts,
                row_index=row_idx
            )
            vendors.append(vendor)
        
        return vendors
    
    def load_master_vendors(self) -> List[VendorRecord]:
        """Load vendors from Master Vendor List."""
        wb = load_workbook(self.master_path, data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        header_map = {}
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                if 'organization' in header_lower or 'vendor' in header_lower:
                    header_map['name'] = i
                elif 'category' in header_lower or 'service' in header_lower:
                    header_map['service_tags'] = i
                elif 'talent' in header_lower or 'contact' in header_lower:
                    header_map['contacts'] = i
                elif 'phone' in header_lower:
                    header_map['phone'] = i
                elif 'email' in header_lower:
                    header_map['email'] = i
                elif 'notes' in header_lower:
                    header_map['notes'] = i
        
        vendors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            name = row[header_map.get('name', 1)].value if header_map.get('name') is not None else None
            if not name or not str(name).strip():
                continue
            
            # Parse service tags (could be comma-separated)
            service_tags = []
            if header_map.get('service_tags') is not None:
                tags_val = row[header_map['service_tags']].value
                if tags_val:
                    tags_str = str(tags_val).strip()
                    service_tags = [t.strip() for t in tags_str.split(',') if t.strip()]
            
            # Parse contacts (could be comma-separated names)
            contacts = []
            if header_map.get('contacts') is not None:
                contacts_val = row[header_map['contacts']].value
                if contacts_val:
                    contacts_str = str(contacts_val).strip()
                    contact_names = [c.strip() for c in contacts_str.split(',') if c.strip()]
                    contacts = [{'name': name} for name in contact_names]
            
            vendor = VendorRecord(
                source='master',
                name=str(name).strip(),
                phone=str(row[header_map.get('phone', 4)].value or "").strip() if header_map.get('phone') is not None else "",
                email=str(row[header_map.get('email', 5)].value or "").strip() if header_map.get('email') is not None else "",
                service_tags=service_tags,
                vendor_notes=str(row[header_map.get('notes', 6)].value or "").strip() if header_map.get('notes') is not None else "",
                contacts=contacts,
                row_index=row_idx
            )
            vendors.append(vendor)
        
        return vendors
    
    def find_exact_matches(self):
        """Phase 1: Find exact matches using normalized name, email, phone."""
        # Group vendors by normalized name
        by_normalized_name: Dict[str, List[VendorRecord]] = defaultdict(list)
        by_email: Dict[str, List[VendorRecord]] = defaultdict(list)
        by_phone: Dict[str, List[VendorRecord]] = defaultdict(list)
        
        for vendor in self.vendors:
            norm_name = vendor.normalize_name()
            norm_email = vendor.normalize_email()
            norm_phone = vendor.normalize_phone()
            
            if norm_name:
                by_normalized_name[norm_name].append(vendor)
            if norm_email:
                by_email[norm_email].append(vendor)
            if norm_phone:
                by_phone[norm_phone].append(vendor)
        
        # Find exact matches
        matched_vendors: List[VendorRecord] = []
        
        # Match by exact normalized name
        for norm_name, vendor_list in by_normalized_name.items():
            if len(vendor_list) > 1:
                match_key = f"name:{norm_name}"
                self.exact_matches[match_key] = vendor_list
                matched_vendors.extend([v for v in vendor_list if v not in matched_vendors])
        
        # Match by email (if not already matched by name)
        for norm_email, vendor_list in by_email.items():
            if len(vendor_list) > 1:
                unmatched = [v for v in vendor_list if v not in matched_vendors]
                if len(unmatched) > 1:
                    match_key = f"email:{norm_email}"
                    if match_key not in self.exact_matches:
                        self.exact_matches[match_key] = []
                    self.exact_matches[match_key].extend(unmatched)
                    matched_vendors.extend([v for v in unmatched if v not in matched_vendors])
        
        # Match by phone (if not already matched)
        for norm_phone, vendor_list in by_phone.items():
            if len(vendor_list) > 1 and norm_phone:
                unmatched = [v for v in vendor_list if v not in matched_vendors]
                if len(unmatched) > 1:
                    match_key = f"phone:{norm_phone}"
                    if match_key not in self.exact_matches:
                        self.exact_matches[match_key] = []
                    self.exact_matches[match_key].extend(unmatched)
                    matched_vendors.extend([v for v in unmatched if v not in matched_vendors])
        
        # Match by name + email or name + phone
        for vendor in self.vendors:
            if vendor in matched_vendors:
                continue
            
            norm_name = vendor.normalize_name()
            norm_email = vendor.normalize_email()
            norm_phone = vendor.normalize_phone()
            
            # Check for name + email match
            if norm_name and norm_email:
                for other in self.vendors:
                    if other == vendor or other in matched_vendors:
                        continue
                    if (other.normalize_name() == norm_name and 
                        other.normalize_email() == norm_email):
                        match_key = f"name+email:{norm_name}:{norm_email}"
                        if match_key not in self.exact_matches:
                            self.exact_matches[match_key] = [vendor]
                        if vendor not in self.exact_matches[match_key]:
                            self.exact_matches[match_key].append(vendor)
                        if other not in self.exact_matches[match_key]:
                            self.exact_matches[match_key].append(other)
                        if vendor not in matched_vendors:
                            matched_vendors.append(vendor)
                        if other not in matched_vendors:
                            matched_vendors.append(other)
            
            # Check for name + phone match
            if norm_name and norm_phone:
                for other in self.vendors:
                    if other == vendor or other in matched_vendors:
                        continue
                    if (other.normalize_name() == norm_name and 
                        other.normalize_phone() == norm_phone):
                        match_key = f"name+phone:{norm_name}:{norm_phone}"
                        if match_key not in self.exact_matches:
                            self.exact_matches[match_key] = [vendor]
                        if vendor not in self.exact_matches[match_key]:
                            self.exact_matches[match_key].append(vendor)
                        if other not in self.exact_matches[match_key]:
                            self.exact_matches[match_key].append(other)
                        if vendor not in matched_vendors:
                            matched_vendors.append(vendor)
                        if other not in matched_vendors:
                            matched_vendors.append(other)
        
        return matched_vendors
    
    def find_fuzzy_matches(self, unmatched_vendors: List[VendorRecord], threshold: float = 0.85):
        """Phase 2: Find fuzzy matches for unmatched vendors."""
        self.fuzzy_matches = []
        processed = []
        
        for i, vendor1 in enumerate(unmatched_vendors):
            if vendor1 in processed:
                continue
            
            best_match = None
            best_score = 0.0
            
            for vendor2 in unmatched_vendors[i+1:]:
                if vendor2 in processed:
                    continue
                
                # Compare normalized names
                score = difflib.SequenceMatcher(
                    None,
                    vendor1.normalize_name(),
                    vendor2.normalize_name()
                ).ratio()
                
                if score >= threshold and score > best_score:
                    best_match = vendor2
                    best_score = score
            
            if best_match:
                self.fuzzy_matches.append((vendor1, best_match, best_score))
                if vendor1 not in processed:
                    processed.append(vendor1)
                if best_match not in processed:
                    processed.append(best_match)
    
    def merge_vendor_group(self, vendor_group: List[VendorRecord]) -> VendorRecord:
        """Merge a group of vendor records into one."""
        # Priority: rippling (base) > qbo (phone) > master (notes/service tags)
        priority_order = {'rippling': 3, 'qbo': 2, 'master': 1}
        sorted_group = sorted(vendor_group, key=lambda v: priority_order.get(v.source, 0), reverse=True)
        
        # Start with Rippling as base, or first vendor if no Rippling
        base_vendor = next((v for v in sorted_group if v.source == 'rippling'), sorted_group[0])
        merged = VendorRecord(
            source='merged',
            name=base_vendor.name,  # Use Rippling name as base
            is_company=True,
            company_name=base_vendor.company_name,
            country=base_vendor.country,
            state=base_vendor.state,
            city=base_vendor.city,
            street=base_vendor.street,
            street2=base_vendor.street2,
            zip=base_vendor.zip,
            email=base_vendor.email,
            contacts=base_vendor.contacts.copy() if base_vendor.contacts else []
        )
        
        # Merge fields with priority rules
        for vendor in sorted_group:
            # Name: prefer Rippling, but use most complete version
            if vendor.source == 'rippling' and vendor.name:
                merged.name = vendor.name
            elif len(vendor.name) > len(merged.name):
                merged.name = vendor.name
            
            # Phone: prefer QBO, then others
            if vendor.source == 'qbo' and vendor.phone:
                merged.phone = vendor.phone
            elif not merged.phone and vendor.phone:
                merged.phone = vendor.phone
            
            # Email: prefer Rippling, then merge others
            if vendor.source == 'rippling' and vendor.email:
                merged.email = vendor.email
            elif vendor.email and vendor.email not in merged.email:
                if merged.email:
                    merged.email += f", {vendor.email}"
                else:
                    merged.email = vendor.email
            
            # Address: prefer Rippling, then fill gaps
            if vendor.source == 'rippling':
                if vendor.street:
                    merged.street = vendor.street
                if vendor.street2:
                    merged.street2 = vendor.street2
                if vendor.city:
                    merged.city = vendor.city
                if vendor.state:
                    merged.state = vendor.state
                if vendor.zip:
                    merged.zip = vendor.zip
                if vendor.country:
                    merged.country = vendor.country
            else:
                # Fill gaps from other sources
                if not merged.street and vendor.street:
                    merged.street = vendor.street
                if not merged.street2 and vendor.street2:
                    merged.street2 = vendor.street2
                if not merged.city and vendor.city:
                    merged.city = vendor.city
                if not merged.state and vendor.state:
                    merged.state = vendor.state
                if not merged.zip and vendor.zip:
                    merged.zip = vendor.zip
                if not merged.country and vendor.country:
                    merged.country = vendor.country
            
            # Company name: prefer Rippling
            if vendor.source == 'rippling' and vendor.company_name:
                merged.company_name = vendor.company_name
            elif not merged.company_name and vendor.company_name:
                merged.company_name = vendor.company_name
            
            # Service tags: prefer master, merge unique
            if vendor.service_tags:
                for tag in vendor.service_tags:
                    if tag not in merged.service_tags:
                        merged.service_tags.append(tag)
            
            # Vendor notes: prefer master
            if not merged.vendor_notes and vendor.vendor_notes:
                merged.vendor_notes = vendor.vendor_notes
            elif vendor.source == 'master' and vendor.vendor_notes:
                merged.vendor_notes = vendor.vendor_notes
            
            # Contacts: prefer Rippling contacts, then merge others
            if vendor.source == 'rippling' and vendor.contacts:
                # Rippling contacts take priority
                merged.contacts = vendor.contacts.copy()
            else:
                # Merge contacts from other sources (avoid duplicates)
                existing_contact_names = {c.get('name', '').lower() for c in merged.contacts}
                for contact in vendor.contacts:
                    contact_name = contact.get('name', '').lower()
                    if contact_name and contact_name not in existing_contact_names:
                        merged.contacts.append(contact)
                        existing_contact_names.add(contact_name)
        
        return merged
    
    def generate_fuzzy_review_file(self):
        """Generate Excel file for reviewing fuzzy matches."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Fuzzy Matches Review"
        
        # Headers
        headers = [
            "Similarity %", "Decision", "Source 1", "Vendor 1 Name", "Vendor 1 Phone", "Vendor 1 Email",
            "Source 2", "Vendor 2 Name", "Vendor 2 Phone", "Vendor 2 Email",
            "Vendor 1 Service Tags", "Vendor 2 Service Tags", "Vendor 1 Notes", "Vendor 2 Notes"
        ]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add fuzzy matches
        for vendor1, vendor2, score in self.fuzzy_matches:
            row = [
                f"{score*100:.1f}%",
                "",  # Decision column (user fills this)
                vendor1.source.upper(),
                vendor1.name,
                vendor1.phone or "",
                vendor1.email or "",
                vendor2.source.upper(),
                vendor2.name,
                vendor2.phone or "",
                vendor2.email or "",
                ", ".join(vendor1.service_tags) if vendor1.service_tags else "",
                ", ".join(vendor2.service_tags) if vendor2.service_tags else "",
                vendor1.vendor_notes[:100] if vendor1.vendor_notes else "",  # Truncate for display
                vendor2.vendor_notes[:100] if vendor2.vendor_notes else ""
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output_path = self.output_dir / "fuzzy_matches_review.xlsx"
        wb.save(output_path)
        print(f"[OK] Fuzzy matches review file saved: {output_path}")
        
        # Also save fuzzy matches metadata for reloading
        fuzzy_data = []
        for vendor1, vendor2, score in self.fuzzy_matches:
            fuzzy_data.append({
                'vendor1_name': vendor1.name,
                'vendor1_source': vendor1.source,
                'vendor2_name': vendor2.name,
                'vendor2_source': vendor2.source,
                'score': score
            })
        metadata_path = self.output_dir / "fuzzy_matches_metadata.json"
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(fuzzy_data, f, indent=2)
        
        return output_path
    
    def load_fuzzy_review_decisions(self) -> Dict[Tuple[str, str], bool]:
        """Load user decisions from fuzzy review file."""
        review_path = self.output_dir / "fuzzy_matches_review.xlsx"
        if not review_path.exists():
            return {}
        
        wb = load_workbook(review_path, data_only=True)
        ws = wb.active
        
        decisions = {}
        for row in ws.iter_rows(min_row=2, values_only=False):
            decision_cell = row[1]  # Decision column (index 1)
            vendor1_name = str(row[3].value or "").strip()  # Vendor 1 Name (index 3)
            vendor2_name = str(row[7].value or "").strip()  # Vendor 2 Name (index 7)
            
            if not vendor1_name or not vendor2_name:
                continue
                
            if decision_cell.value:
                decision_str = str(decision_cell.value).strip().lower()
                should_merge = decision_str in ['merge', 'yes', 'y', '1', 'true']
                decisions[(vendor1_name, vendor2_name)] = should_merge
        
        return decisions
    
    def generate_output_files(self):
        """Generate final merged output files."""
        # Load vendor template to get column order
        wb_template = load_workbook(self.vendor_template_path, data_only=True)
        ws_template = wb_template.active
        vendor_headers = [str(cell.value or "").strip() for cell in ws_template[1] if cell.value]
        
        # Add custom vendor fields if not present in template
        if not any('x_vendor_notes' in str(h).lower() for h in vendor_headers):
            vendor_headers.append('x_vendor_notes')
        if not any('x_vendor_service_tag_ids' in str(h).lower() or ('service' in str(h).lower() and 'tag' in str(h).lower()) for h in vendor_headers):
            vendor_headers.append('x_vendor_service_tag_ids')
        # Add contacts/talent column to show people associated with each company
        if not any('contact' in str(h).lower() and 'talent' in str(h).lower() or 'x_vendor_contacts' in str(h).lower() for h in vendor_headers):
            vendor_headers.append('x_vendor_contacts')
        # Add external ID for reliable linking (if not present)
        if not any('id' in str(h).lower() and ('external' in str(h).lower() or 'xml' in str(h).lower()) for h in vendor_headers):
            vendor_headers.append('id')  # External ID field
        
        # Create merged vendors workbook
        wb_vendors = Workbook()
        ws_vendors = wb_vendors.active
        ws_vendors.append(vendor_headers)
        
        # Style header
        for cell in ws_vendors[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add merged vendor records
        for idx, vendor in enumerate(self.merged_vendors, start=1):
            row_data = {}
            # Generate external ID for vendor (for reliable contact linking)
            vendor_external_id = f"vendor_{idx:04d}"
            
            for header in vendor_headers:
                header_lower = str(header).lower().strip()
                if header_lower == 'name':
                    row_data[header] = vendor.name
                elif header_lower == 'is_company':
                    row_data[header] = 'True' if vendor.is_company else 'False'
                elif header_lower == 'company_name':
                    row_data[header] = vendor.company_name
                elif 'country' in header_lower:
                    row_data[header] = vendor.country  # User will need to map to country ID in Odoo
                elif 'state' in header_lower:
                    row_data[header] = vendor.state  # User will need to map to state ID in Odoo
                elif header_lower == 'zip':
                    row_data[header] = vendor.zip
                elif header_lower == 'city':
                    row_data[header] = vendor.city
                elif header_lower == 'street' and '2' not in header_lower:
                    row_data[header] = vendor.street
                elif 'street2' in header_lower or ('street' in header_lower and '2' in header_lower):
                    row_data[header] = vendor.street2
                elif header_lower == 'phone':
                    row_data[header] = vendor.phone
                elif header_lower == 'mobile':
                    row_data[header] = ""  # Not in source data
                elif header_lower == 'email':
                    row_data[header] = vendor.email.split(',')[0] if vendor.email else ""  # Take first email
                elif 'vat' in header_lower or 'tax' in header_lower:
                    row_data[header] = ""  # Not in source data
                elif 'bank' in header_lower:
                    row_data[header] = ""  # Not in source data
                elif 'x_vendor_service_tag_ids' in header_lower or 'service' in header_lower:
                    # Service tags - comma separated for import
                    row_data[header] = ", ".join(vendor.service_tags) if vendor.service_tags else ""
                elif 'x_vendor_notes' in header_lower or ('notes' in header_lower and 'vendor' in header_lower):
                    row_data[header] = vendor.vendor_notes
                elif 'x_vendor_contacts' in header_lower or ('contact' in header_lower and 'talent' in header_lower):
                    # Show all contacts/talent associated with this vendor company
                    contact_names = [c.get('name', '') for c in vendor.contacts if c.get('name')]
                    row_data[header] = ", ".join(contact_names) if contact_names else ""
                elif header_lower == 'id' and 'external' not in header_lower and 'xml' not in header_lower:
                    # External ID for reliable linking
                    row_data[header] = vendor_external_id
            
            row = [str(row_data.get(header, "") or "") for header in vendor_headers]
            ws_vendors.append(row)
        
        # Auto-adjust column widths
        for column in ws_vendors.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_vendors.column_dimensions[column_letter].width = adjusted_width
        
        # Load contact template to get the combined format
        wb_contact_template = load_workbook(self.contact_template_path, data_only=True)
        ws_contact_template = wb_contact_template.active
        combined_headers = [str(cell.value or "").strip() for cell in ws_contact_template[1] if cell.value]
        
        # Create ONE combined file with companies and contacts
        wb_combined = Workbook()
        ws_combined = wb_combined.active
        ws_combined.append(combined_headers)
        
        # Style header
        for cell in ws_combined[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add companies and their contacts together
        for idx, vendor in enumerate(self.merged_vendors, start=1):
            vendor_external_id = f"vendor_{idx:04d}"
            
            # First, add the COMPANY row
            company_row_data = {}
            for header in combined_headers:
                header_lower = str(header).lower().strip()
                header_str = str(header)
                
                # Name* (required) - company name
                if 'name' in header_lower and '*' in header_str:
                    company_row_data[header] = vendor.name
                # Company Type* (required) - "Company" for companies
                elif 'company type' in header_lower and '*' in header_str:
                    company_row_data[header] = 'Company'
                # Related Company - empty for companies
                elif 'related company' in header_lower:
                    company_row_data[header] = ""
                # Email
                elif 'email' in header_lower:
                    company_row_data[header] = vendor.email.split(',')[0] if vendor.email else ""
                # Phone
                elif 'phone' in header_lower:
                    company_row_data[header] = vendor.phone
                # Address fields
                elif 'street' in header_lower and '2' not in header_lower:
                    company_row_data[header] = vendor.street
                elif 'street2' in header_lower or ('street' in header_lower and '2' in header_lower):
                    company_row_data[header] = vendor.street2
                elif 'city' in header_lower:
                    company_row_data[header] = vendor.city
                elif 'state' in header_lower:
                    company_row_data[header] = vendor.state
                elif 'zip' in header_lower:
                    company_row_data[header] = vendor.zip
                elif 'country' in header_lower:
                    company_row_data[header] = vendor.country
                # Notes - use vendor notes
                elif 'notes' in header_lower:
                    company_row_data[header] = vendor.vendor_notes
                # Tags - use service tags
                elif 'tags' in header_lower:
                    company_row_data[header] = ", ".join(vendor.service_tags) if vendor.service_tags else ""
                # Reference - use external ID
                elif 'reference' in header_lower:
                    company_row_data[header] = vendor_external_id
                # Other fields
                else:
                    company_row_data[header] = ""
            
            company_row = [str(company_row_data.get(header, "") or "") for header in combined_headers]
            ws_combined.append(company_row)
            
            # Then, add CONTACT rows beneath the company
            if vendor.contacts:
                for contact_num, contact in enumerate(vendor.contacts, start=1):
                    contact_row_data = {}
                    for header in combined_headers:
                        header_lower = str(header).lower().strip()
                        header_str = str(header)
                        
                        # Name* (required) - contact name
                        if 'name' in header_lower and '*' in header_str:
                            contact_row_data[header] = contact.get('name', '')
                        # Company Type* (required) - "Person" for contacts
                        elif 'company type' in header_lower and '*' in header_str:
                            contact_row_data[header] = 'Person'
                        # Related Company - link to parent vendor by name
                        elif 'related company' in header_lower:
                            contact_row_data[header] = vendor.name
                        # Email
                        elif 'email' in header_lower:
                            contact_row_data[header] = contact.get('email', vendor.email.split(',')[0] if vendor.email else "")
                        # Phone
                        elif 'phone' in header_lower:
                            contact_row_data[header] = contact.get('phone', vendor.phone)
                        # Address fields - LEAVE EMPTY for contacts (they inherit from company)
                        elif 'street' in header_lower and '2' not in header_lower:
                            contact_row_data[header] = ""  # Contacts inherit address from company
                        elif 'street2' in header_lower or ('street' in header_lower and '2' in header_lower):
                            contact_row_data[header] = ""  # Contacts inherit address from company
                        elif 'city' in header_lower:
                            contact_row_data[header] = ""  # Contacts inherit address from company
                        elif 'state' in header_lower:
                            contact_row_data[header] = ""  # Contacts inherit address from company
                        elif 'zip' in header_lower:
                            contact_row_data[header] = ""  # Contacts inherit address from company
                        elif 'country' in header_lower:
                            contact_row_data[header] = ""  # Contacts inherit address from company
                        # Notes
                        elif 'notes' in header_lower:
                            contact_row_data[header] = f"Contact for {vendor.name}"
                        # Reference - link to company's external ID (numbered per vendor, starting at 1)
                        elif 'reference' in header_lower:
                            contact_row_data[header] = f"{vendor_external_id}_contact_{contact_num}"
                        # Other fields
                        else:
                            contact_row_data[header] = ""
                    
                    contact_row = [str(contact_row_data.get(header, "") or "") for header in combined_headers]
                    ws_combined.append(contact_row)
        
        # Auto-adjust column widths
        for column in ws_combined.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_combined.column_dimensions[column_letter].width = adjusted_width
        
        # Count total contacts
        total_contacts = sum(len(v.contacts) for v in self.merged_vendors)
        
        # Save combined file
        combined_path = self.output_dir / "merged_vendors_and_contacts.xlsx"
        wb_combined.save(combined_path)
        print(f"[OK] Combined vendors and contacts file saved: {combined_path}")
        print(f"     - {len(self.merged_vendors)} companies")
        print(f"     - {total_contacts} contacts")
        
        # Also save separate files for backward compatibility
        vendors_path = self.output_dir / "merged_vendors.xlsx"
        wb_vendors.save(vendors_path)
        print(f"[OK] Separate vendors file saved: {vendors_path}")
    
    def generate_reports(self):
        """Generate merge reports."""
        report_lines = []
        report_lines.append("=" * 80)
        report_lines.append("VENDOR DATA MERGE REPORT")
        report_lines.append("=" * 80)
        report_lines.append("")
        
        # Summary
        report_lines.append(f"Total vendors loaded: {len(self.vendors)}")
        report_lines.append(f"  - QBO: {sum(1 for v in self.vendors if v.source == 'qbo')}")
        report_lines.append(f"  - Rippling: {sum(1 for v in self.vendors if v.source == 'rippling')}")
        report_lines.append(f"  - Master: {sum(1 for v in self.vendors if v.source == 'master')}")
        report_lines.append("")
        
        # Exact matches
        report_lines.append(f"Exact matches found: {len(self.exact_matches)}")
        for match_key, vendor_group in self.exact_matches.items():
            report_lines.append(f"  {match_key}: {len(vendor_group)} vendors")
            for vendor in vendor_group:
                report_lines.append(f"    - {vendor.name} ({vendor.source})")
        report_lines.append("")
        
        # Fuzzy matches
        report_lines.append(f"Fuzzy matches found: {len(self.fuzzy_matches)}")
        for vendor1, vendor2, score in self.fuzzy_matches:
            report_lines.append(f"  {score*100:.1f}%: '{vendor1.name}' ({vendor1.source}) <-> '{vendor2.name}' ({vendor2.source})")
        report_lines.append("")
        
        # Final merged count
        report_lines.append(f"Final merged vendors: {len(self.merged_vendors)}")
        report_lines.append("")
        
        report_text = "\n".join(report_lines)
        
        # Save report
        report_path = self.output_dir / "merge_report.txt"
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report_text)
        
        print("\n" + report_text)
        print(f"\n✓ Merge report saved: {report_path}")
    
    def run(self, fuzzy_threshold: float = 0.85):
        """Run the complete merge process."""
        print("Loading vendor data from source files...")
        
        # Load all vendors
        qbo_vendors = self.load_qbo_vendors()
        rippling_vendors = self.load_rippling_vendors()
        master_vendors = self.load_master_vendors()
        
        self.vendors = qbo_vendors + rippling_vendors + master_vendors
        print(f"[OK] Loaded {len(self.vendors)} total vendor records")
        
        # Phase 1: Exact matching
        print("\nPhase 1: Finding exact matches...")
        matched_vendors = self.find_exact_matches()
        print(f"[OK] Found {len(self.exact_matches)} exact match groups")
        
        # Merge exact matches
        exact_matched_list = []
        for match_key, vendor_group in self.exact_matches.items():
            merged = self.merge_vendor_group(vendor_group)
            self.merged_vendors.append(merged)
            exact_matched_list.extend([v for v in vendor_group if v not in exact_matched_list])
        
        # Phase 2: Fuzzy matching
        print("\nPhase 2: Finding fuzzy matches...")
        unmatched_vendors = [v for v in self.vendors if v not in exact_matched_list]
        self.find_fuzzy_matches(unmatched_vendors, threshold=fuzzy_threshold)
        print(f"[OK] Found {len(self.fuzzy_matches)} fuzzy matches")
        
        # Check if review file exists
        review_path = self.output_dir / "fuzzy_matches_review.xlsx"
        review_exists = review_path.exists()
        
        # Generate fuzzy review file if needed
        if self.fuzzy_matches and not review_exists:
            print("\nGenerating fuzzy matches review file...")
            self.generate_fuzzy_review_file()
            print("\n[!] Please review fuzzy_matches_review.xlsx and mark 'Merge' or 'Keep Separate' in the Decision column.")
            print("   Then run this script again to apply your decisions.")
            return
        
        # Phase 3: Apply fuzzy decisions and finalize
        if review_exists:
            print("\nPhase 3: Applying merge decisions...")
            fuzzy_decisions = self.load_fuzzy_review_decisions()
            
            # Reconstruct fuzzy matches from review file if needed
            # (in case vendors were reordered or fuzzy matching changed)
            if not self.fuzzy_matches:
                # Try to reload from metadata or reconstruct from review file
                metadata_path = self.output_dir / "fuzzy_matches_metadata.json"
                if metadata_path.exists():
                    with open(metadata_path, 'r', encoding='utf-8') as f:
                        fuzzy_data = json.load(f)
                    
                    # Rebuild fuzzy matches by finding vendors by name
                    vendor_by_name_source = {}
                    for vendor in self.vendors:
                        key = (vendor.name, vendor.source)
                        vendor_by_name_source[key] = vendor
                    
                    for match_data in fuzzy_data:
                        v1_key = (match_data['vendor1_name'], match_data['vendor1_source'])
                        v2_key = (match_data['vendor2_name'], match_data['vendor2_source'])
                        if v1_key in vendor_by_name_source and v2_key in vendor_by_name_source:
                            vendor1 = vendor_by_name_source[v1_key]
                            vendor2 = vendor_by_name_source[v2_key]
                            if vendor1 not in exact_matched_list and vendor2 not in exact_matched_list:
                                self.fuzzy_matches.append((vendor1, vendor2, match_data['score']))
            
            # Merge approved fuzzy matches
            processed_fuzzy = []
            for vendor1, vendor2, score in self.fuzzy_matches:
                key1 = (str(vendor1.name), str(vendor2.name))
                key2 = (str(vendor2.name), str(vendor1.name))
                
                should_merge = False
                if key1 in fuzzy_decisions:
                    should_merge = fuzzy_decisions[key1]
                elif key2 in fuzzy_decisions:
                    should_merge = fuzzy_decisions[key2]
                
                if should_merge and vendor1 not in processed_fuzzy and vendor2 not in processed_fuzzy:
                    merged = self.merge_vendor_group([vendor1, vendor2])
                    self.merged_vendors.append(merged)
                    if vendor1 not in processed_fuzzy:
                        processed_fuzzy.append(vendor1)
                    if vendor2 not in processed_fuzzy:
                        processed_fuzzy.append(vendor2)
            
            # Add unmatched vendors (not in exact matches, not merged in fuzzy)
            for vendor in unmatched_vendors:
                if vendor not in processed_fuzzy:
                    self.merged_vendors.append(vendor)
        else:
            # No fuzzy matches or no review file - add all unmatched vendors
            for vendor in unmatched_vendors:
                self.merged_vendors.append(vendor)
        
        # CRITICAL: Rebuild merged_vendors starting from Rippling as the base
        # Master Vendor List and QBO are only for adding data, not for adding new vendors
        print("\nRebuilding vendor list from Rippling base...")
        
        # Normalize function for matching
        def normalize_for_matching(name):
            """Normalize name for better matching."""
            if not name:
                return ""
            name = str(name).strip().lower()
            # Remove common suffixes
            name = re.sub(r'\s+(inc|llc|ltd|corp|corporation|company|co)\.?$', '', name, flags=re.IGNORECASE)
            # Remove extra whitespace and punctuation
            name = re.sub(r'[^\w\s]', '', name)
            name = re.sub(r'\s+', ' ', name)
            return name.strip()
        
        # Create lookup maps for QBO and Master vendors (multiple strategies)
        qbo_by_exact = {v.name.lower().strip(): v for v in qbo_vendors if v.name}
        qbo_by_normalized = {normalize_for_matching(v.name): v for v in qbo_vendors if v.name}
        
        # Debug: Check QBO vendors with phones
        qbo_with_phones = [v for v in qbo_vendors if v.phone and v.phone.strip()]
        print(f"[DEBUG] QBO vendors loaded: {len(qbo_vendors)}, with phones: {len(qbo_with_phones)}")
        if qbo_with_phones:
            print(f"[DEBUG] Sample QBO vendor with phone: '{qbo_with_phones[0].name}' -> '{qbo_with_phones[0].phone}'")
        
        master_by_exact = {v.name.lower().strip(): v for v in master_vendors if v.name}
        master_by_normalized = {normalize_for_matching(v.name): v for v in master_vendors if v.name}
        
        # Debug: Check Master vendors with notes
        master_with_notes = [v for v in master_vendors if v.vendor_notes and v.vendor_notes.strip()]
        print(f"[DEBUG] Master vendors loaded: {len(master_vendors)}, with notes: {len(master_with_notes)}")
        
        # Also create fuzzy matching maps (for high-confidence matches)
        from difflib import SequenceMatcher
        qbo_fuzzy_map = {}  # Will be built on-demand
        master_fuzzy_map = {}  # Will be built on-demand
        
        # Start with Rippling vendors and enrich them
        # Use a dict to deduplicate by normalized name (case-insensitive)
        final_merged_vendors_dict = {}  # normalized_name -> VendorRecord
        qbo_matched = 0
        master_matched = 0
        qbo_fuzzy_matched = 0
        master_fuzzy_matched = 0
        
        for rippling_vendor in rippling_vendors:
            # Use company_name if available, otherwise name
            rippling_name = (rippling_vendor.company_name or rippling_vendor.name or "").strip()
            rippling_name_lower = rippling_name.lower()
            rippling_name_normalized = normalize_for_matching(rippling_name)
            
            # Start with Rippling vendor as base
            merged_vendor = VendorRecord(
                source='merged',
                name=rippling_vendor.name,
                is_company=True,
                company_name=rippling_vendor.company_name,
                country=rippling_vendor.country,
                state=rippling_vendor.state,
                city=rippling_vendor.city,
                street=rippling_vendor.street,
                street2=rippling_vendor.street2,
                zip=rippling_vendor.zip,
                email=rippling_vendor.email,
                contacts=rippling_vendor.contacts.copy() if rippling_vendor.contacts else []
            )
            
            # Enrich with QBO data (phone numbers) - try exact, normalized, then fuzzy
            qbo_vendor = None
            if rippling_name_lower in qbo_by_exact:
                qbo_vendor = qbo_by_exact[rippling_name_lower]
            elif rippling_name_normalized in qbo_by_normalized:
                qbo_vendor = qbo_by_normalized[rippling_name_normalized]
            else:
                # Try fuzzy matching (similarity > 0.80 for better coverage, especially for phone numbers)
                best_match = None
                best_score = 0.0
                for q_vendor in qbo_vendors:
                    if not q_vendor.name:
                        continue
                    q_norm = normalize_for_matching(q_vendor.name)
                    if q_norm:
                        score = SequenceMatcher(None, rippling_name_normalized, q_norm).ratio()
                        # Lower threshold to 0.80 to catch more matches (especially for phones)
                        if score > best_score and score > 0.80:
                            best_score = score
                            best_match = q_vendor
                if best_match:
                    qbo_vendor = best_match
                    qbo_fuzzy_matched += 1
            
            # Debug first few matches
            if qbo_matched < 3 and qbo_vendor:
                print(f"[DEBUG] Matched Rippling '{rippling_name}' -> QBO '{qbo_vendor.name}' (phone: '{qbo_vendor.phone}')")
            
            if qbo_vendor:
                qbo_matched += 1
                # Always use QBO phone if available (QBO is the source for phones)
                if qbo_vendor.phone and qbo_vendor.phone.strip():
                    merged_vendor.phone = qbo_vendor.phone
                # Also merge other QBO data if missing
                if not merged_vendor.street and qbo_vendor.street:
                    merged_vendor.street = qbo_vendor.street
                # Use QBO email if Rippling doesn't have one, or if QBO email is more complete
                if qbo_vendor.email and qbo_vendor.email.strip():
                    if not merged_vendor.email or not merged_vendor.email.strip():
                        merged_vendor.email = qbo_vendor.email
                    # If both have emails, prefer QBO if it looks more complete (has @)
                    elif '@' in qbo_vendor.email and '@' not in merged_vendor.email:
                        merged_vendor.email = qbo_vendor.email
            
            # Enrich with Master Vendor List data (service tags and notes) - try exact, normalized, then fuzzy
            master_vendor = None
            if rippling_name_lower in master_by_exact:
                master_vendor = master_by_exact[rippling_name_lower]
            elif rippling_name_normalized in master_by_normalized:
                master_vendor = master_by_normalized[rippling_name_normalized]
            else:
                # Try fuzzy matching (similarity > 0.80 for better coverage)
                best_match = None
                best_score = 0.0
                for m_vendor in master_vendors:
                    if not m_vendor.name:
                        continue
                    m_norm = normalize_for_matching(m_vendor.name)
                    if m_norm:
                        score = SequenceMatcher(None, rippling_name_normalized, m_norm).ratio()
                        if score > best_score and score > 0.80:  # Lowered from 0.85
                            best_score = score
                            best_match = m_vendor
                if best_match:
                    master_vendor = best_match
                    master_fuzzy_matched += 1
            
            if master_vendor:
                master_matched += 1
                if master_vendor.service_tags:
                    merged_vendor.service_tags = master_vendor.service_tags.copy()
                if master_vendor.vendor_notes:
                    merged_vendor.vendor_notes = master_vendor.vendor_notes
                # Merge contacts from Master if any
                if master_vendor.contacts:
                    existing_names = {c.get('name', '').lower() for c in merged_vendor.contacts}
                    for contact in master_vendor.contacts:
                        if contact.get('name', '').lower() not in existing_names:
                            merged_vendor.contacts.append(contact)
            
            # Check if we already have a vendor with this normalized name (case-insensitive duplicate)
            if rippling_name_normalized in final_merged_vendors_dict:
                # Merge with existing vendor instead of creating duplicate
                existing_vendor = final_merged_vendors_dict[rippling_name_normalized]
                
                # Prefer proper case name over all caps or all lowercase
                def is_better_case(name1, name2):
                    """Determine which name has better casing (prefer Title Case over ALL CAPS or all lowercase)."""
                    if not name1 or not name2:
                        return name1 or name2
                    # Prefer name that's not all caps and not all lowercase
                    name1_is_all_caps = name1.isupper()
                    name2_is_all_caps = name2.isupper()
                    name1_is_all_lower = name1.islower()
                    name2_is_all_lower = name2.islower()
                    
                    # Prefer Title Case over all caps or all lower
                    if name1_is_all_caps and not name2_is_all_caps:
                        return name2
                    if name2_is_all_caps and not name1_is_all_caps:
                        return name1
                    if name1_is_all_lower and not name2_is_all_lower:
                        return name2
                    if name2_is_all_lower and not name1_is_all_lower:
                        return name1
                    # If both are same case quality, prefer the one that's not empty
                    return name1 if name1 else name2
                
                # Update name to best case version
                existing_vendor.name = is_better_case(existing_vendor.name, merged_vendor.name)
                existing_vendor.company_name = is_better_case(existing_vendor.company_name, merged_vendor.company_name)
                
                # Merge phone (prefer non-empty)
                if merged_vendor.phone and merged_vendor.phone.strip() and not existing_vendor.phone:
                    existing_vendor.phone = merged_vendor.phone
                
                # Merge email (prefer non-empty)
                if merged_vendor.email and merged_vendor.email.strip() and not existing_vendor.email:
                    existing_vendor.email = merged_vendor.email
                
                # Merge address fields (prefer non-empty)
                if merged_vendor.street and not existing_vendor.street:
                    existing_vendor.street = merged_vendor.street
                if merged_vendor.street2 and not existing_vendor.street2:
                    existing_vendor.street2 = merged_vendor.street2
                if merged_vendor.city and not existing_vendor.city:
                    existing_vendor.city = merged_vendor.city
                if merged_vendor.state and not existing_vendor.state:
                    existing_vendor.state = merged_vendor.state
                if merged_vendor.zip and not existing_vendor.zip:
                    existing_vendor.zip = merged_vendor.zip
                if merged_vendor.country and not existing_vendor.country:
                    existing_vendor.country = merged_vendor.country
                
                # Merge service tags (combine unique)
                if merged_vendor.service_tags:
                    existing_tags = set(existing_vendor.service_tags or [])
                    for tag in merged_vendor.service_tags:
                        if tag not in existing_tags:
                            existing_vendor.service_tags.append(tag)
                
                # Merge notes (prefer longer/more complete)
                if merged_vendor.vendor_notes and len(merged_vendor.vendor_notes) > len(existing_vendor.vendor_notes or ""):
                    existing_vendor.vendor_notes = merged_vendor.vendor_notes
                
                # Merge contacts (avoid duplicates)
                if merged_vendor.contacts:
                    existing_contact_names = {c.get('name', '').lower() for c in existing_vendor.contacts}
                    for contact in merged_vendor.contacts:
                        if contact.get('name', '').lower() not in existing_contact_names:
                            existing_vendor.contacts.append(contact)
            else:
                # New vendor, add it
                final_merged_vendors_dict[rippling_name_normalized] = merged_vendor
        
        # Convert dict to list
        final_merged_vendors = list(final_merged_vendors_dict.values())
        
        print(f"[OK] Matched {qbo_matched} vendors with QBO ({qbo_matched - qbo_fuzzy_matched} exact, {qbo_fuzzy_matched} fuzzy)")
        print(f"[OK] Matched {master_matched} vendors with Master ({master_matched - master_fuzzy_matched} exact, {master_fuzzy_matched} fuzzy)")
        print(f"[OK] Deduplicated {len(rippling_vendors) - len(final_merged_vendors)} case-sensitive duplicates")
        
        # Replace merged_vendors with Rippling-based list
        original_count = len(self.merged_vendors)
        self.merged_vendors = final_merged_vendors
        print(f"[OK] Rebuilt vendor list: {len(self.merged_vendors)} vendors (all from Rippling, enriched with QBO/Master data)")
        
        # Generate output files
        print("\nGenerating output files...")
        self.generate_output_files()
        
        # Generate reports
        print("\nGenerating reports...")
        self.generate_reports()
        
        print("\n✓ Merge complete!")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Merge vendor data from multiple sources")
    parser.add_argument('--qbo', default=r'C:\Users\ashpt\Downloads\Copy of Vendors QBO.xlsx',
                       help='Path to QBO vendors file')
    parser.add_argument('--rippling', default=r'C:\Users\ashpt\Downloads\Copy of Vendors Rippling.xlsx',
                       help='Path to Rippling vendors file')
    parser.add_argument('--master', default=r'C:\Users\ashpt\Downloads\Master Vendors List (1).xlsx',
                       help='Path to Master Vendor List file')
    parser.add_argument('--vendor-template', default=r'C:\Users\ashpt\Downloads\vendors_import_res_partner (2).xlsx',
                       help='Path to vendor import template')
    parser.add_argument('--contact-template', default=r'C:\Users\ashpt\Downloads\contacts_import_template.xlsx',
                       help='Path to contact import template')
    parser.add_argument('--output-dir', default='.',
                       help='Output directory for merged files')
    parser.add_argument('--fuzzy-threshold', type=float, default=0.85,
                       help='Fuzzy matching similarity threshold (0.0-1.0)')
    
    args = parser.parse_args()
    
    # Validate input files
    for path_name, path_value in [
        ('QBO', args.qbo),
        ('Rippling', args.rippling),
        ('Master', args.master),
        ('Vendor Template', args.vendor_template),
        ('Contact Template', args.contact_template)
    ]:
        if not Path(path_value).exists():
            print(f"[ERROR] {path_name} file not found: {path_value}")
            sys.exit(1)
    
    # Create output directory
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Run merger
    merger = VendorMerger(
        qbo_path=args.qbo,
        rippling_path=args.rippling,
        master_path=args.master,
        vendor_template_path=args.vendor_template,
        contact_template_path=args.contact_template,
        output_dir=output_dir
    )
    
    try:
        merger.run(fuzzy_threshold=args.fuzzy_threshold)
    except Exception as e:
        print(f"\n[ERROR] Error during merge: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
